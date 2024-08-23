import openpyxl
from openpyxl import load_workbook
from xlwt import Workbook
from xlsxwriter.workbook import Workbook
from co_occurance_mitomap import query_mitomap_covariants, parse_covariant_data


""" convert var format for search in genomAD (ex 152TC -> M-152-T-C) """
def convert_var_format(var):
    var = var.value
    position = var[:-2]
    nucleotide1 = var[-2]
    nucleotide2 = var[-1]
    var_format = ['M-', position, '-', nucleotide1, '-', nucleotide2]

    return "".join(var_format)


""" extract all variants from a given excel file and create
 variants_dict = {var:[patients with this var]}
  and patients_dict = {patient:[all variants this patient have]} """
def extract_all_variants(path):
    # To open the workbook - workbook object is created
    wb_obj = openpyxl.load_workbook(path)

    # Get workbook active sheet object from the active attribute
    sheet_obj = wb_obj.active

    rows = sheet_obj.max_row
    columns = sheet_obj.max_column
    #print("Total Rows:", rows, "Total Columns:", columns)

    variants_dict = {}
    patients_dict = {}
    variants_set = set()

    for column in range(1, columns+1):
        patient = sheet_obj.cell(1, column)
        patients_dict[patient.value] = []
        for row in range(2, rows + 1):
            cell_obj = sheet_obj.cell(row, column)
            if not cell_obj.value:
                break

            cell_obj.value = convert_var_format(cell_obj)
            if cell_obj.value not in variants_dict:
                variants_dict[cell_obj.value] = []

            patients_dict[patient.value].append(cell_obj.value)
            variants_dict[cell_obj.value].append(patient.value)
            variants_set.add(cell_obj.value)

    #print("num of uniqe variants =", len(variants_set))
    #print(variants_set)

    return variants_dict, patients_dict


""" Write variants_dict = {var:[patients with this var]} to a new excel sheet """
def write_new_file(variants_dict):
    wb = Workbook()
    sheet1 = wb.add_sheet('Sheet 1')
    sheet1.write(0, 0, "variant")
    sheet1.write(0, 1, "patients_with_var")

    row = 1
    for var in variants_dict:
        sheet1.write(row, 0, var)
        patients_with_var = ", ".join(variants_dict[var])
        sheet1.write(row, 1, patients_with_var)
        row += 1

    wb.save('all_different_variants1.xls')
    print("done!")


""" Create very_common_variants dict containing all the patients missing one variant out of 6 most common variants
     ('M-263-A-G', 'M-750-A-G', 'M-1438-A-G', 'M-4769-A-G', 'M-8860-A-G', 'M-15326-A-G') """
def very_common_var(patients_dict, very_common_variants):
    for patient in patients_dict:
        for var in very_common_variants:
            if var not in patients_dict[patient]:
                very_common_variants[var].append(patient.value)

    return very_common_variants


""" given a patient number and the patients dict (output of the func "extract_all_variants") -
   this func will create a list with all the variants this patient have (exclude the most common ones)
    and then call the function "query_mitomap_covariants" which create a report (excel) with the relevant data from mitomap co-ocuuarnce tool """
def analyse_patient_mitomap(patient_num, input_patients_dict):
    print("\ncase " + patient_num + ":")
    patient_variants_lst = input_patients_dict.get(patient_num)
    unique_variants = []
    for var in patient_variants_lst:
        if var not in very_common_variants:
            var = var.split('-')
            mitomap_foramt = var[1] + var[3]
            unique_variants.append(mitomap_foramt)

    res = query_mitomap_covariants(unique_variants)
    parse_analyse_patient_res(patient_num, res)


def parse_analyse_patient_res(patient_num, result):
    df = parse_covariant_data(result)
    file_name = f"case_{patient_num}_mitomap_frequencies.xlsx"
    df = df.set_index("Variant")
    df.to_excel(file_name)

    # Write to Excel using openpyxl
    wb = load_workbook(file_name)
    ws = wb.active

    # Autofit columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_name)

    print(f"Excel file '{file_name}' has been created and formatted.")


if __name__ == '__main__':
    # path to the variants excel file
    path = r"C:\Users\sheer\Documents\לימודים\אוניבסיטת תא\טביעת אצבע\all_patients_variantsALL2.xlsx"
    variants_dict, patients_dict = extract_all_variants(path)
    #write_new_file(variants_dict)

    very_common_variants = {'M-263-A-G': [], 'M-750-A-G': [], 'M-1438-A-G': [], 'M-4769-A-G': [], 'M-8860-A-G': [], 'M-15326-A-G': []}
    #print("\n missing very_common_variants\n", very_common_var(patients_dict, very_common_variants))


    """ given an input patient number, get an excel file with all the variants showed count smaller then 500 (out of 61134 in the database),
     and all the frequencies for these variants according to mitomap"""
    #analyse_patient_mitomap("70X", patients_dict)
    #analyse_patient_mitomap("47X", patients_dict)
    #analyse_patient_mitomap("91X", patients_dict)
    #analyse_patient_mitomap("94Y", patients_dict)
    #analyse_patient_mitomap("191", patients_dict)
    #analyse_patient_mitomap("62X", patients_dict)
    #analyse_patient_mitomap("25X", patients_dict)
    #analyse_patient_mitomap("129Y", patients_dict)
    #analyse_patient_mitomap("42Y", patients_dict)
    #analyse_patient_mitomap("79X", patients_dict)
    #analyse_patient_mitomap("140", patients_dict)
    #analyse_patient_mitomap("125Y", patients_dict)
